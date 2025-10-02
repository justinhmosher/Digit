from __future__ import annotations

import json
from io import BytesIO
from typing import Optional

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect
from django.views.decorators.http import require_GET, require_POST

from .models import (
    OwnerProfile,
    RestaurantProfile,
    Ownership,
    ManagerProfile,   # assumes you have this as in your manager views
    TicketLink,
    OwnerInvite,
    ManagerInvite,
    StaffInvite,
    StaffProfile,
    Review
)
from .omnivore import get_ticket, get_ticket_items
from django.views.decorators.http import require_http_methods
from django.db import transaction
User = get_user_model()
from .utils import send_manager_invite_email, send_owner_invite_email, send_staff_invite_email
import re

# ----------------- helpers -----------------

# DROP-IN: replace your existing owner_api_state with this
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_GET
from django.utils import timezone
from django.http import JsonResponse, HttpRequest
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.apps import apps


def _get_owner_profile(user):
    op = getattr(user, "owner_profile", None)
    if op is None:
        op = OwnerProfile.objects.select_related("user").filter(user=user).first()
    return op


def _owner_restaurants(op):
    # Restaurants the owner has access to via Ownership
    return (
        RestaurantProfile.objects
        .filter(ownerships__owner=op)
        .distinct()
    )


def _get_current_restaurant(request, op):
    qs = _owner_restaurants(op)
    if not qs.exists():
        return None

    rid = request.session.get("current_restaurant_id")
    current = qs.filter(id=rid).first() if rid else None
    if not current:
        current = qs.first()

    # persist
    if request.session.get("current_restaurant_id") != current.id:
        request.session["current_restaurant_id"] = current.id
        request.session.modified = True
    return current


@ensure_csrf_cookie
@require_GET
@login_required
def owner_api_state(request: HttpRequest) -> JsonResponse:
    """
    Owner dashboard data:
      - restaurants list + current selection
      - owners (emails)
      - managers (names)
      - staff (names)
      - open tickets summary
      - recent closed tickets with optional filters
    """
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)

    # Restaurants this owner controls
    rqs = _owner_restaurants(op).order_by("created_at")
    restaurants = [
        {
            "id": r.id,
            "name": r.dba_name or r.legal_name,
            "phone": r.phone or "",
            "email": r.email or "",
        }
        for r in rqs
    ]

    current = _get_current_restaurant(request, op)
    current_id = current.id if current else None

    # Query params for orders
    q = (request.GET.get("q") or "").strip().lower()
    start = (request.GET.get("start") or "").strip()
    end   = (request.GET.get("end") or "").strip()

    # Owners (via Ownership -> OwnerProfile.user.email)
    owners = []
    if current:
        links = (
            Ownership.objects
            .select_related("owner", "owner__user")
            .filter(restaurant=current)
        )
        for lk in links:
            u = getattr(lk.owner, "user", None)
            owners.append({"id": lk.owner_id, "email": (u.email if u else "")})

    # Managers (names)
    managers = []
    if current:
        mqs = (
            ManagerProfile.objects
            .select_related("user")
            .filter(restaurant=current)
            .order_by("user__email")
        )
        for m in mqs:
            u = getattr(m, "user", None)
            email = (u.email if u else "")
            name = (u.get_full_name() if u else "") or (email.split("@")[0] if email else "")
            managers.append({"id": m.id, "name": name})

    # Staff (names)
    staff = []
    if current:
        from .models import StaffProfile  # ensure available
        sqs = (
            StaffProfile.objects
            .select_related("user")
            .filter(restaurant=current)
            .order_by("user__email")
        )
        for s in sqs:
            u = getattr(s, "user", None)
            email = (u.email if u else "")
            name = (u.get_full_name() if u else "") or (email.split("@")[0] if email else "")
            staff.append({"id": s.id, "name": name})

    # Open tickets (group by ticket_id)
    open_map = {}
    if current:
        open_qs = (
            TicketLink.objects
            .select_related("member")
            .filter(restaurant=current, status="open")
            .order_by("-opened_at")[:400]
        )
        for tl in open_qs:
            entry = open_map.setdefault(tl.ticket_id, {
                "ticket_id": tl.ticket_id,
                "ticket_number": tl.ticket_number or None,
                "server": tl.server_name or "",
                "members": [],
                "due_cents": 0,
            })
            entry["members"].append(tl.member.number if tl.member else "")
            entry["due_cents"] = max(entry["due_cents"], tl.last_total_cents or 0)
    open_list = list(open_map.values())

    # Recent closed
    recent = []
    if current:
        recent_qs = TicketLink.objects.select_related("member").filter(restaurant=current, status="closed")
        if start:
            try:
                recent_qs = recent_qs.filter(closed_at__date__gte=start)
            except Exception:
                pass
        if end:
            try:
                recent_qs = recent_qs.filter(closed_at__date__lte=end)
            except Exception:
                pass

        for tl in recent_qs.order_by("-closed_at")[:500]:
            row = {
                "ticket_id": tl.ticket_id,
                "ticket_number": tl.ticket_number or None,
                "member": tl.member.number if tl.member else "",
                "server": tl.server_name or "",
                "closed_at": tl.closed_at.strftime("%Y-%m-%d %H:%M") if tl.closed_at else "",
                "total_cents": (tl.paid_cents or tl.total_cents or tl.last_total_cents or 0),
                "ticket_link_id": tl.id,  # NEW: used by Review button
            }
            if q:
                hay = f'{row["ticket_number"] or ""} {row["member"] or ""}'.lower()
                if q not in hay:
                    continue
            recent.append(row)

    return JsonResponse({
        "ok": True,
        "restaurants": restaurants,
        "current_restaurant_id": current_id,
        "owners": owners,
        "managers": managers,
        "staff": staff,
        "open": open_list,
        "recent": recent,
    })




@login_required
@require_http_methods(["GET", "POST"])
def owner_invite_manager(request):
    """Owner sends an invite for the *current* restaurant."""
    rp = _current_restaurant(request)
    if request.method == "GET":
        # If you prefer to block here, you can render a page explaining they must onboard first.
        if not rp:
            return JsonResponse({"ok": False, "error": "Create your restaurant profile first."}, status=400)
        return render(request, "core/owner_invite_manager.html", {"restaurant": rp})

    # POST (JSON or form)
    if not rp:
        return JsonResponse({"ok": False, "error": "Create your restaurant profile first."}, status=400)

    email = ""
    expires_minutes = 120

    # JSON body first
    try:
        payload = json.loads((request.body or b"").decode() or "{}")
        email = (payload.get("email") or "").strip().lower()
        if payload.get("expires_minutes") is not None:
            expires_minutes = int(payload["expires_minutes"])
    except Exception:
        pass

    # Fallback to form POST
    if not email:
        email = (request.POST.get("email") or "").strip().lower()
    if request.POST.get("expires_minutes"):
        try:
            expires_minutes = int(request.POST.get("expires_minutes"))
        except ValueError:
            pass

    if not email:
        return JsonResponse({"ok": False, "error": "Please provide an email."}, status=400)

    invite = ManagerInvite.objects.create(
        restaurant=rp,
        email=email,
        expires_at=timezone.now() + timedelta(minutes=expires_minutes),
    )

    link = f"{request.scheme}://{request.get_host()}/manager/accept?token={invite.token}"
    rest_name = rp.dba_name or rp.legal_name or "your restaurant"

    try:
        send_manager_invite_email(
            to_email=email,
            invite_link=link,
            restaurant_name=rest_name,
            expires_at=invite.expires_at,
        )
        email_ok = True
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Email send failed: {e}"}, status=500)

    return JsonResponse({
        "ok": True,
        "message": f"Invite sent to {email}.",
        "invite": {
            "email": email,
            "token": str(invite.token),
            "expires_at": invite.expires_at.isoformat(),
            "link": link,
            "email_sent": email_ok,
        },
    })

def _set_current_restaurant(request: HttpRequest, rp: Optional[RestaurantProfile]):
    request.session["current_restaurant_id"] = rp.id if rp else None
    request.session.modified = True

# ----------------- page -----------------
def _current_restaurant(request):
    """Resolve the restaurant for the signed-in owner."""
    # 1) session selection
    rid = request.session.get("current_restaurant_id")
    if rid:
        rp = RestaurantProfile.objects.filter(id=rid).first()
        if rp:
            return rp

    # 2) first active restaurant this owner owns
    op = OwnerProfile.objects.filter(user=request.user).first()
    if not op:
        return None

    # Use through model; respect is_active if present
    ow_qs = Ownership.objects.filter(owner=op)
    if any(f.name == "is_active" for f in Ownership._meta.fields):
        ow_qs = ow_qs.filter(is_active=True)

    rid = ow_qs.values_list("restaurant_id", flat=True).first()
    if rid:
        rp = RestaurantProfile.objects.filter(id=rid).first()
        if rp:
            # remember it in session for next time
            request.session["current_restaurant_id"] = rp.id
            request.session.modified = True
            return rp
    return None

@login_required
def owner_dashboard(request: HttpRequest) -> HttpResponse:
    op = _get_owner_profile(request.user)
    if not op:
        # If the logged-in user doesn't have an OwnerProfile yet, bail with a clear error.
        return render(request, "core/owner_dashboard.html", {"op": None, "restaurant": None})
    rp = _get_current_restaurant(request, op)
    _set_current_restaurant(request, rp)
    return render(request, "core/owner_dashboard.html", {"op": op, "restaurant": rp})

# ----------------- state -----------------

# ----------------- mutations -----------------

@csrf_protect
@require_POST
@login_required
def owner_api_set_restaurant(request: HttpRequest) -> JsonResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}
    rid = data.get("restaurant_id")
    rp = _owner_restaurants(op).filter(id=rid).first()
    if not rp:
        return JsonResponse({"ok": False, "error": "Restaurant not found."}, status=404)
    _set_current_restaurant(request, rp)
    return JsonResponse({"ok": True})

@csrf_protect
@require_POST
@login_required
def owner_api_add_restaurant(request: HttpRequest) -> JsonResponse:
    """
    Create a restaurant and link it to the current owner via Ownership.
    Your model requires legal_name and email.
    """
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}
    legal = (data.get("legal_name") or "").strip()
    dba   = (data.get("dba_name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    if not legal or not email:
        return JsonResponse({"ok": False, "error": "Legal name and Email are required."}, status=400)

    rp = RestaurantProfile.objects.create(
        legal_name=legal,
        dba_name=dba,
        email=email,
        phone=phone or "",
    )
    # link ownership
    Ownership.objects.get_or_create(owner=op, restaurant=rp)
    _set_current_restaurant(request, rp)
    return JsonResponse({"ok": True, "restaurant_id": rp.id})

# DROP-IN: hard delete a restaurant the owner controls
@csrf_protect
@require_POST
@login_required
def owner_api_remove_restaurant(request: HttpRequest) -> JsonResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)

    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}

    rid = data.get("restaurant_id")
    if not rid:
        return JsonResponse({"ok": False, "error": "Missing restaurant_id."}, status=400)

    # Owner must own this restaurant
    rp = _owner_restaurants(op).filter(id=rid).first()
    if not rp:
        return JsonResponse({"ok": False, "error": "Restaurant not found."}, status=404)

    with transaction.atomic():
        # Keep user accounts, just unlink them from this restaurant
        StaffProfile.objects.filter(restaurant=rp).update(restaurant=None)
        ManagerProfile.objects.filter(restaurant=rp).update(restaurant=None)

        # TicketLink has FK(PROTECT) -> must be deleted first or deletion is blocked
        TicketLink.objects.filter(restaurant=rp).delete()

        deleted_id = rp.id
        rp.delete()  # cascades Ownership, invites, etc.

        # Reset session selection to another restaurant if available
        request.session.pop("current_restaurant_id", None)
        next_r = _owner_restaurants(op).order_by("created_at").first()
        if next_r:
            request.session["current_restaurant_id"] = next_r.id
        request.session.modified = True

    return JsonResponse({
        "ok": True,
        "deleted_id": deleted_id,
        "current_restaurant_id": request.session.get("current_restaurant_id")
    })

@csrf_protect
@require_POST
@login_required
def owner_api_remove_manager(request: HttpRequest) -> JsonResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}
    rid = data.get("restaurant_id")
    mid = data.get("manager_id")
    rp = _owner_restaurants(op).filter(id=rid).first()
    if not rp:
        return JsonResponse({"ok": False, "error": "Restaurant not found."}, status=404)

    mp = ManagerProfile.objects.filter(id=mid, restaurant=rp).first()
    if not mp:
        return JsonResponse({"ok": False, "error": "Manager not found."}, status=404)

    mp.restaurant = None
    mp.save(update_fields=["restaurant"])
    return JsonResponse({"ok": True})

@login_required
@require_http_methods(["GET", "POST"])
def owner_api_add_owner(request):
    """Owner sends an invite for the *current* restaurant."""
    rp = _current_restaurant(request)
    if request.method == "GET":
        # If you prefer to block here, you can render a page explaining they must onboard first.
        if not rp:
            return JsonResponse({"ok": False, "error": "Create your restaurant profile first."}, status=400)
        return render(request, "core/owner_invite_owbner.html", {"restaurant": rp})

    # POST (JSON or form)
    if not rp:
        return JsonResponse({"ok": False, "error": "Create your restaurant profile first."}, status=400)

    email = ""
    expires_minutes = 120

    # JSON body first
    try:
        payload = json.loads((request.body or b"").decode() or "{}")
        email = (payload.get("email") or "").strip().lower()
        if payload.get("expires_minutes") is not None:
            expires_minutes = int(payload["expires_minutes"])
    except Exception:
        pass

    # Fallback to form POST
    if not email:
        email = (request.POST.get("email") or "").strip().lower()
    if request.POST.get("expires_minutes"):
        try:
            expires_minutes = int(request.POST.get("expires_minutes"))
        except ValueError:
            pass

    if not email:
        return JsonResponse({"ok": False, "error": "Please provide an email."}, status=400)

    invite = OwnerInvite.objects.create(
        restaurant=rp,
        email=email,
        expires_at=timezone.now() + timedelta(minutes=expires_minutes),
    )

    link = f"{request.scheme}://{request.get_host()}/owner/accept?token={invite.token}"
    rest_name = rp.dba_name or rp.legal_name or "your restaurant"

    try:
        send_owner_invite_email(
            to_email=email,
            invite_link=link,
            restaurant_name=rest_name,
            expires_at=invite.expires_at,
        )
        email_ok = True
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Email send failed: {e}"}, status=500)

    return JsonResponse({
        "ok": True,
        "message": f"Invite sent to {email}.",
        "invite": {
            "email": email,
            "token": str(invite.token),
            "expires_at": invite.expires_at.isoformat(),
            "link": link,
            "email_sent": email_ok,
        },
    })

def _set_current_restaurant(request: HttpRequest, rp: Optional[RestaurantProfile]):
    request.session["current_restaurant_id"] = rp.id if rp else None
    request.session.modified = True

@csrf_protect
@require_POST
@login_required
def owner_api_remove_owner(request: HttpRequest) -> JsonResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}
    rid = data.get("restaurant_id")
    oid = data.get("owner_id")
    rp = _owner_restaurants(op).filter(id=rid).first()
    if not rp:
        return JsonResponse({"ok": False, "error": "Restaurant not found."}, status=404)

    # Don’t allow removing last owner
    total = rp.owners.count()
    if total <= 1:
        return JsonResponse({"ok": False, "error": "Cannot remove the last owner."}, status=400)

    Ownership.objects.filter(owner_id=oid, restaurant=rp).delete()
    return JsonResponse({"ok": True})

# ----------------- receipts / export -----------------

@require_GET
@login_required
def owner_api_ticket_detail(request: HttpRequest, ticket_id: str) -> JsonResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    rp = _get_current_restaurant(request, op)
    if not rp:
        return JsonResponse({"ok": False, "error": "Select a restaurant."}, status=400)

    tl = (
        TicketLink.objects
        .filter(restaurant=rp, ticket_id=str(ticket_id))
        .order_by("-opened_at")
        .first()
    )
    if not tl:
        return JsonResponse({"ok": False, "error": "Ticket not found."}, status=404)

    # OPEN (live pull): Subtotal = Total display; include per-item unit price
    if tl.status == "open" and (rp.omnivore_location_id or "").strip():
        try:
            t = get_ticket(rp.omnivore_location_id, tl.ticket_id)
            items = get_ticket_items(rp.omnivore_location_id, tl.ticket_id)
        except Exception as e:
            return JsonResponse({"ok": False, "error": f"POS error: {e}"}, status=502)

        rows = []
        for it in items:
            qty   = int(it.get("quantity", 1) or 1)
            unit  = int(it.get("price", 0) or 0)   # per-unit price in cents from Omnivore
            line  = unit * qty
            rows.append({
                "name": it.get("name"),
                "qty": qty,
                "unit_cents": unit,
                "line_total_cents": line,
                "cents": unit,  # back-compat for older UI
            })

        totals = (t or {}).get("totals") or {}
        tax = int(totals.get("tax", 0) or 0)
        tip = int(totals.get("tip", 0) or 0)

        try:
            due_val = totals.get("due")
            total = int(due_val) if due_val is not None else int(totals.get("total", 0) or 0)
        except Exception:
            total = 0
        if total <= 0:
            line_sum = sum(int(r["line_total_cents"]) for r in rows)
            total = line_sum + tax + tip

        return JsonResponse({
            "ok": True,
            "ticket_id": tl.ticket_id,
            "ticket_number": tl.ticket_number or (t.get("ticket_number") or t.get("number") or tl.ticket_id),
            "member": tl.member.number if tl.member else "",
            "server": tl.server_name or ((t.get("_embedded") or {}).get("employee") or {}).get("check_name", ""),
            "items": rows,
            "subtotal_cents": total,  # display rule
            "tax_cents": tax,
            "tip_cents": tip,
            "total_cents": total,
            "is_open": True,
        })

    # CLOSED: use snapshot captured at close; expose unit + line totals
    rows = []
    for it in (tl.items_json or []):
        name  = it.get("name") or it.get("label") or "Item"
        qty   = int(it.get("qty") or it.get("quantity") or 1)
        unit  = int(it.get("price_cents") or it.get("unit_cents") or it.get("cents") or it.get("price") or 0)
        line  = int(it.get("total_cents") or it.get("line_total_cents") or (unit * qty))
        rows.append({
            "name": name,
            "qty": qty,
            "unit_cents": unit,
            "line_total_cents": line,
            "cents": unit,  # back-compat
        })

    return JsonResponse({
        "ok": True,
        "ticket_id": tl.ticket_id,
        "ticket_number": tl.ticket_number or tl.ticket_id,
        "member": tl.member.number if tl.member else "",
        "server": tl.server_name or "",
        "items": rows,
        "subtotal_cents": int(tl.total_cents or 0),  # per your mapping
        "tax_cents": int(tl.tax_cents or 0),
        "tip_cents": int(tl.tip_cents or 0),
        "total_cents": int(tl.paid_cents or (tl.total_cents or 0) + (tl.tax_cents or 0) + (tl.tip_cents or 0)),
        "is_open": False,
    })


def _owner_restaurant_or_404(request):
    # however you currently get the active restaurant (reusing your helper if you have one)
    op = _get_owner_profile(request.user)
    if not op:
        raise Http404("Owner not found")
    rp = _get_current_restaurant(request, op)
    if not rp:
        raise Http404("Restaurant not selected")
    return rp

@login_required
def owner_ticket_review_json(request, ticket_link_id: int) -> JsonResponse:
    rp = _owner_restaurant_or_404(request)

    try:
        tl = TicketLink.objects.select_related("member").get(
            pk=ticket_link_id, restaurant=rp
        )
    except TicketLink.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Ticket not found."}, status=404)

    # If you store one review per ticket, this is fine; otherwise pick your rule
    review = (
        Review.objects.filter(ticket_link=tl)
        .order_by("-created_at")
        .first()
    )

    out = {
        "ok": True,
        "ticket_id": tl.ticket_id,
        "ticket_number": tl.ticket_number,
        "member": getattr(tl.member, "member_id", None) or getattr(tl.member, "id", None),
        "has_review": bool(review),
        "review": None,
    }

    if review:
        # Stars must be a number 0..5
        stars = review.stars or 0  # adjust field name if different
        try:
            stars = int(stars)
        except Exception:
            try:
                stars = float(stars)
            except Exception:
                stars = 0
        stars = max(0, min(5, stars))

        out["review"] = {
            "id": review.id,
            "stars": stars,     # canonical
            "rating": stars,    # redundant alias for front-end safety
            "comment": review.comment or "",
            "created_at": review.created_at.isoformat() if review.created_at else None,
            "reviewer_name": getattr(review, "reviewer_name", "") or "",
            "reviewer_email": getattr(review, "reviewer_email", "") or "",
        }

    return JsonResponse(out)
# core/utils_reviews.py
from django.apps import apps

# Common attribute names across projects
_POSSIBLE_RATING_ATTRS = ("rating", "review_rating", "stars", "score")
# Common JSON/blob fields that may contain nested review data
_POSSIBLE_REVIEW_FIELDS = ("review_json", "raw_review_json", "raw_ticket_json", "extra_json")

def _dig_rating_from_mapping(m):
    """
    Try common keys/paths inside a dict-like review/ticket payload.
    Return int 0–5 (usually 1–5) or None.
    """
    if not isinstance(m, dict):
        return None

    # Direct keys first
    for k in ("rating", "review_rating", "stars", "score"):
        try:
            v = m.get(k)
            if v is not None:
                v = int(v)
                if 0 <= v <= 5:
                    return v
        except Exception:
            pass

    # Nested "review": { rating: ... }
    try:
        rev = m.get("review") or {}
        v = rev.get("rating")
        if v is not None:
            v = int(v)
            if 0 <= v <= 5:
                return v
    except Exception:
        pass

    # Other common nesting patterns
    for path in (("details", "rating"), ("feedback", "rating"), ("customer", "rating")):
        try:
            cur = m
            for k in path:
                if not isinstance(cur, dict):
                    cur = {}
                cur = cur.get(k)
            if cur is not None:
                v = int(cur)
                if 0 <= v <= 5:
                    return v
        except Exception:
            pass

    return None


# core/utils_reviews.py
from django.apps import apps

_POSSIBLE_RATING_ATTRS = ("rating", "review_rating", "stars", "score")
_POSSIBLE_REVIEW_FIELDS = ("review_json", "raw_review_json", "raw_ticket_json", "extra_json")

def _dig_rating_from_mapping(m):
    if not isinstance(m, dict):
        return None
    for k in ("rating", "review_rating", "stars", "score"):
        try:
            v = m.get(k)
            if v is not None:
                v = int(v)
                if 0 <= v <= 5:
                    return v
        except Exception:
            pass
    try:
        rev = m.get("review") or {}
        v = rev.get("rating")
        if v is not None:
            v = int(v)
            if 0 <= v <= 5:
                return v
    except Exception:
        pass
    for path in (("details", "rating"), ("feedback", "rating"), ("customer", "rating")):
        try:
            cur = m
            for k in path:
                if not isinstance(cur, dict):
                    cur = {}
                cur = cur.get(k)
            if cur is not None:
                v = int(cur)
                if 0 <= v <= 5:
                    return v
        except Exception:
            pass
    return None

def get_ticket_rating_from_anywhere(ticket_link):
    # 1) Review models
    for label in ("core.Review", "core.TicketReview", "reviews.Review"):
        try:
            Review = apps.get_model(label)
        except Exception:
            Review = None
        if not Review:
            continue

        try:
            r = Review.objects.filter(ticket_link=ticket_link).order_by("-id").first()
        except Exception:
            r = None

        if r is None:
            try:
                filt = {"ticket_id": ticket_link.ticket_id}
                if hasattr(Review, "restaurant"):
                    filt["restaurant"] = ticket_link.restaurant
                r = Review.objects.filter(**filt).order_by("-id").first()
            except Exception:
                r = None

        if r is not None:
            for attr in _POSSIBLE_RATING_ATTRS:
                try:
                    val = getattr(r, attr, None)
                    if val is not None:
                        val = int(val)
                        if 0 <= val <= 5:
                            return val
                except Exception:
                    pass
            for blob_attr in _POSSIBLE_REVIEW_FIELDS:
                try:
                    blob = getattr(r, blob_attr, None)
                    if blob:
                        v = _dig_rating_from_mapping(blob)
                        if v is not None:
                            return v
                except Exception:
                    pass

    # 2) Direct attrs on TicketLink
    for attr in _POSSIBLE_RATING_ATTRS:
        try:
            val = getattr(ticket_link, attr, None)
            if val is not None:
                val = int(val)
                if 0 <= val <= 5:
                    return val
        except Exception:
            pass

    # 3) Dict/JSON blobs on TicketLink
    for blob_attr in _POSSIBLE_REVIEW_FIELDS:
        try:
            blob = getattr(ticket_link, blob_attr, None)
            if blob:
                v = _dig_rating_from_mapping(blob)
                if v is not None:
                    return v
        except Exception:
            pass

    return None


# --- MENU ITEMS ANALYTICS ---
@login_required
@require_GET
def owner_api_menu_item_ratings(request: HttpRequest) -> JsonResponse:
    """
    Menu analytics that ALWAYS returns rows from closed tickets:
      - If a rating exists on a ticket, include it in the average.
      - If not, show the item with avg_rating=None and volume counts.
      - Price comes from RestaurantProfile.menu_cache, else ticket unit price.
    """

    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    rp = _get_current_restaurant(request, op)
    if not rp:
        return JsonResponse({"ok": False, "error": "Select a restaurant."}, status=400)

    start_s = (request.GET.get("start") or "").strip()
    end_s   = (request.GET.get("end") or "").strip()
    qs = TicketLink.objects.filter(restaurant=rp, status="closed")
    if start_s:
        try: qs = qs.filter(closed_at__date__gte=parse_date(start_s))
        except Exception: pass
    if end_s:
        try: qs = qs.filter(closed_at__date__lte=parse_date(end_s))
        except Exception: pass

    item_meta = { str(x.get("id")): x for x in (rp.menu_cache or []) }

    agg = {}
    for tl in qs.iterator():
        rating = get_ticket_rating_from_anywhere(tl)  # may be None

        for row in (tl.items_json or []):
            mid = str(row.get("menu_item_id") or row.get("id") or "").strip()
            name = (row.get("name") or row.get("label") or "").strip() or "Unknown item"
            key = mid or f"name:{name}"

            try:
                qty = int(row.get("quantity") or row.get("qty") or 1)
            except Exception:
                qty = 1

            rec = agg.setdefault(key, {
                "menu_item_id": mid or None,
                "name": name,
                "category": "",
                "price_cents": 0,
                "sum": 0,
                "n": 0,
                "qty_rated_tickets": 0,
                "qty_all_tickets": 0,
                "num_rated_tickets": 0,
                "num_all_tickets": 0,
            })

            # enrich from cache
            if mid and mid in item_meta:
                meta = item_meta[mid]
                if not rec["category"]:
                    rec["category"] = meta.get("category") or ""
                try:
                    if not rec["price_cents"]:
                        rec["price_cents"] = int(meta.get("price_cents") or 0)
                except Exception:
                    pass

            # fallback to ticket unit price if cache has none
            if not rec["price_cents"]:
                for k in ("unit_cents", "price_cents", "cents", "unit_price", "price"):
                    v = row.get(k)
                    if v is None: 
                        continue
                    try:
                        iv = int(v)
                        rec["price_cents"] = iv if iv >= 100 else iv * 100
                        break
                    except Exception:
                        try:
                            from decimal import Decimal
                            rec["price_cents"] = int(Decimal(str(v)) * 100)
                            break
                        except Exception:
                            pass

            rec["qty_all_tickets"] += qty
            rec["num_all_tickets"] += 1

            if rating is not None:
                rec["sum"] += int(rating)
                rec["n"]   += 1
                rec["qty_rated_tickets"] += qty
                rec["num_rated_tickets"] += 1

    out = []
    for _, r in agg.items():
        avg = (r["sum"] / r["n"]) if r["n"] else None
        out.append({
            "menu_item_id": r["menu_item_id"] or "",
            "name": r["name"],
            "category": r["category"],
            "price_cents": int(r["price_cents"] or 0),
            "avg_rating": round(avg, 3) if avg is not None else None,
            "num_rated_tickets": r["num_rated_tickets"],
            "total_qty_on_rated_tickets": r["qty_rated_tickets"],
            # extra (optional) fields if your UI wants them later:
            "num_all_tickets": r["num_all_tickets"],
            "total_qty_all_tickets": r["qty_all_tickets"],
        })

    def sort_key(row):
        rated = row["avg_rating"] is not None
        return (0 if rated else 1, -(row["avg_rating"] or 0), -row.get("num_all_tickets", 0))

    out.sort(key=sort_key)
    return JsonResponse({"ok": True, "items": out})


# views.py
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from django.http import JsonResponse, HttpRequest

@login_required
@require_GET
def owner_api_staff_ratings(request: HttpRequest) -> JsonResponse:
    """
    Staff analytics that ALWAYS returns rows (but no 'Unknown').
    """

    # --- access ---
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    rp = _get_current_restaurant(request, op)
    if not rp:
        return JsonResponse({"ok": False, "error": "Select a restaurant."}, status=400)

    # --- filters ---
    start_s = (request.GET.get("start") or "").strip()
    end_s   = (request.GET.get("end") or "").strip()
    qs = TicketLink.objects.filter(restaurant=rp, status="closed")
    if start_s:
        try: qs = qs.filter(closed_at__date__gte=parse_date(start_s))
        except Exception: pass
    if end_s:
        try: qs = qs.filter(closed_at__date__lte=parse_date(end_s))
        except Exception: pass

    # --- cache maps ---
    staff_cache = rp.staff_cache or []
    by_ck_lower = { (s.get("check_name") or "").strip().lower(): s for s in staff_cache if (s.get("check_name") or "").strip() }
    by_nm_lower = { (s.get("name") or "").strip().lower():       s for s in staff_cache if (s.get("name") or "").strip() }

    # --- seed rows + buckets ---
    agg = {}
    def seed_row(key, display, active=True):
        if key not in agg:
            agg[key] = {
                "display": display or "",
                "active": bool(active),
                "sum": 0, "n": 0,
                "tickets_all": 0, "tickets_rated": 0
            }

    # seed every cached staffer so they appear even with 0 tickets
    for s in staff_cache:
        key = (str(s.get("id") or "").strip()
               or (s.get("check_name") or "").strip()
               or (s.get("name") or "").strip()
               or f"seed:{id(s)}")
        disp = (s.get("check_name") or s.get("name") or "").strip()
        if disp:  # ✅ only seed if we actually have a display name
            seed_row(key, disp, s.get("is_active", True))

    agg_all = {"display": "All staff", "active": True, "sum": 0, "n": 0, "tickets_all": 0, "tickets_rated": 0}

    # --- aggregation ---
    for tl in qs.iterator():
        rating = get_ticket_rating_from_anywhere(tl)

        # ALL row
        agg_all["tickets_all"] += 1
        if rating is not None:
            agg_all["sum"] += int(rating)
            agg_all["n"]   += 1
            agg_all["tickets_rated"] += 1

        resolved = (tl.server_name or "").strip()
        key = None
        display = None

        if resolved:
            low = resolved.lower()
            if low in by_ck_lower:
                s = by_ck_lower[low]
                key = str(s.get("id") or s.get("check_name") or resolved)
                display = (s.get("check_name") or s.get("name") or resolved)
                seed_row(key, display, s.get("is_active", True))
            elif low in by_nm_lower:
                s = by_nm_lower[low]
                key = str(s.get("id") or s.get("name") or resolved)
                display = (s.get("check_name") or s.get("name") or resolved)
                seed_row(key, display, s.get("is_active", True))
            else:
                # keep raw server_name, but skip if it's blank
                display = resolved
                key = resolved
                if display:
                    seed_row(key, display, True)

        if not key:
            try:
                raw = tl.raw_ticket_json or {}
                emb = (raw.get("_embedded") or {})
                emp = emb.get("employee") or raw.get("employee") or {}
                pos_name = (emp.get("check_name") or emp.get("name") or "").strip()
            except Exception:
                pos_name = ""
            if pos_name:
                low = pos_name.lower()
                if low in by_ck_lower:
                    s = by_ck_lower[low]
                    key = str(s.get("id") or s.get("check_name") or pos_name)
                    display = (s.get("check_name") or s.get("name") or pos_name)
                    seed_row(key, display, s.get("is_active", True))
                elif low in by_nm_lower:
                    s = by_nm_lower[low]
                    key = str(s.get("id") or s.get("name") or pos_name)
                    display = (s.get("check_name") or s.get("name") or pos_name)
                    seed_row(key, display, s.get("is_active", True))
                else:
                    display = pos_name
                    key = pos_name
                    if display:
                        seed_row(key, display, True)

        # ✅ If still no key, SKIP (don’t create “Unknown”)
        if not key:
            continue

        row = agg[key]
        row["tickets_all"] += 1
        if rating is not None:
            row["sum"] += int(rating)
            row["n"]   += 1
            row["tickets_rated"] += 1

    # --- shape response ---
    out = []
    avg_all = (agg_all["sum"] / agg_all["n"]) if agg_all["n"] else None
    out.append({
        "staff_key": "ALL",
        "name": agg_all["display"],
        "avg_rating": round(avg_all, 3) if avg_all is not None else None,
        "num_rated_tickets": agg_all["tickets_rated"],
        "is_active_in_pos": True,
        "num_all_tickets": agg_all["tickets_all"],
    })
    for key, r in agg.items():
        avg = (r["sum"] / r["n"]) if r["n"] else None
        out.append({
            "staff_key": key,
            "name": r["display"],
            "avg_rating": round(avg, 3) if avg is not None else None,
            "num_rated_tickets": r["tickets_rated"],
            "is_active_in_pos": bool(r["active"]),
            "num_all_tickets": r["tickets_all"],
        })

    def sort_key(row):
        if row["staff_key"] == "ALL": return (-999, 0, 0)
        rated = row["avg_rating"] is not None
        return (0 if rated else 1, -(row["avg_rating"] or 0), -row["num_all_tickets"])
    out[1:] = sorted(out[1:], key=sort_key)

    return JsonResponse({
        "ok": True,
        "synced_at": (rp.staff_cache_synced_at.isoformat() if rp.staff_cache_synced_at else None),
        "staff": out,
    })


# views.py
@login_required
@require_GET
def owner_api_staff_ratings_debug(request: HttpRequest) -> JsonResponse:
    """Quick visibility into distinct server names & how many tickets have them."""
    from django.utils.dateparse import parse_date
    from collections import Counter
    from .models import TicketLink

    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    rp = _get_current_restaurant(request, op)
    if not rp:
        return JsonResponse({"ok": False, "error": "Select a restaurant."}, status=400)

    start_s = (request.GET.get("start") or "").strip()
    end_s   = (request.GET.get("end") or "").strip()
    qs = TicketLink.objects.filter(restaurant=rp, status="closed")
    if start_s:
        try: qs = qs.filter(closed_at__date__gte=parse_date(start_s))
        except Exception: pass
    if end_s:
        try: qs = qs.filter(closed_at__date__lte=parse_date(end_s))
        except Exception: pass

    names = []
    for tl in qs.iterator():
        nm = (tl.server_name or "").strip()
        if not nm:
            try:
                raw = tl.raw_ticket_json or {}
                emb = (raw.get("_embedded") or {})
                emp = emb.get("employee") or raw.get("employee") or {}
                nm = (emp.get("check_name") or emp.get("name") or "").strip()
            except Exception:
                nm = ""
        names.append(nm or "<blank>")

    counts = Counter(names)
    rows = [{"server_name": k, "tickets": v} for k, v in sorted(counts.items(), key=lambda x: (-x[1], x[0]))]
    return JsonResponse({"ok": True, "rows": rows})


@require_GET
@login_required
def owner_export(request: HttpRequest) -> HttpResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return HttpResponse("Not an owner.", status=403)

    q = (request.GET.get("q") or "").strip().lower()
    start = (request.GET.get("start") or "").strip()
    end   = (request.GET.get("end") or "").strip()

    rp = _get_current_restaurant(request, op)
    if not rp:
        return HttpResponse("Select a restaurant.", status=400)

    qs = TicketLink.objects.select_related("member").filter(restaurant=rp, status="closed")
    if start:
        try: qs = qs.filter(closed_at__date__gte=start)
        except Exception: pass
    if end:
        try: qs = qs.filter(closed_at__date__lte=end)
        except Exception: pass

    rows = []
    for tl in qs.order_by("closed_at"):
        if q:
            hay = f"{(tl.ticket_number or tl.ticket_id)} {(tl.member.number if tl.member else '')}".lower()
            if q not in hay: continue

        subtotal = int(tl.total_cents or 0)   # per your mapping
        tax      = int(tl.tax_cents or 0)
        tip      = int(tl.tip_cents or 0)
        total    = int(tl.paid_cents or (subtotal + tax + tip))

        rows.append({
            "closed": tl.closed_at,
            "ticket": tl.ticket_number or tl.ticket_id,
            "member": tl.member.number if tl.member else "",
            "server": tl.server_name or "",
            "subtotal": subtotal,
            "tax": tax,
            "tip": tip,
            "total": total,
            "pos_ref": tl.pos_ref or "",
        })

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Closed"

    headers = ["Closed", "Ticket", "Member", "Server", "Subtotal", "Tax", "Tip", "Total", "POS Ref"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    currency_fmt = u'[$$-409]#,##0.00'

    for r in rows:
        ws.append([
            r["closed"].strftime("%Y-%m-%d %H:%M") if r["closed"] else "",
            r["ticket"], r["member"], r["server"],
            r["subtotal"]/100.0, r["tax"]/100.0, r["tip"]/100.0, r["total"]/100.0,
            r["pos_ref"],
        ])

    if rows:
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=8):
            for cell in row:
                cell.number_format = currency_fmt

    widths = [18, 12, 14, 12, 12, 12, 12, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=7, value="Grand total").font = Font(bold=True)
    total_col = get_column_letter(8)
    ws.cell(row=total_row, column=8, value=f"=SUM({total_col}2:{total_col}{total_row-1})").number_format = currency_fmt
    ws.cell(row=total_row, column=8).font = Font(bold=True)

    rest_name = _display_name(rp)
    filename = f"{rest_name.replace(' ', '_')}_closed_{timezone.now().strftime('%Y%m%d')}.xlsx"

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

@csrf_protect
@require_POST
@login_required
def owner_api_remove_staff(request: HttpRequest) -> JsonResponse:
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)
    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}
    rid = data.get("restaurant_id")
    sid = data.get("staff_id")
    rp = _owner_restaurants(op).filter(id=rid).first()
    if not rp:
        return JsonResponse({"ok": False, "error": "Restaurant not found."}, status=404)

    sp = StaffProfile.objects.filter(id=sid, restaurant=rp).first()
    if not sp:
        return JsonResponse({"ok": False, "error": "Staff not found."}, status=404)

    sp.restaurant = None  # unlink access
    sp.save(update_fields=["restaurant"])
    return JsonResponse({"ok": True})

from datetime import timedelta

@csrf_protect
@require_POST
@login_required
def owner_invite_staff(request: HttpRequest) -> JsonResponse:
    """Owner sends a staff invite for the current restaurant."""
    op = _get_owner_profile(request.user)
    if not op:
        return JsonResponse({"ok": False, "error": "Not an owner."}, status=403)

    rp = _get_current_restaurant(request, op)
    if not rp:
        return JsonResponse({"ok": False, "error": "Create/select a restaurant first."}, status=400)

    try:
        payload = json.loads((request.body or b"").decode() or "{}")
    except Exception:
        payload = {}
    email = (payload.get("email") or "").strip().lower()
    expires_minutes = int(payload.get("expires_minutes") or 7*24*60)

    if not email:
        return JsonResponse({"ok": False, "error": "Please provide an email."}, status=400)

    # Create invite (model mirrors ManagerInvite)
    invite = StaffInvite.objects.create(
        restaurant=rp,
        email=email,
        expires_at=timezone.now() + timedelta(minutes=expires_minutes),
    )

    link = f"{request.scheme}://{request.get_host()}/staff/accept?token={invite.token}"
    rest_name = rp.dba_name or rp.legal_name or "your restaurant"

    try:
        send_staff_invite_email(
            to_email=email,
            invite_link=link,
            restaurant_name=rest_name,
            expires_at=invite.expires_at,
        )
        email_ok = True
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Email send failed: {e}"}, status=500)

    return JsonResponse({
        "ok": True,
        "message": f"Invite sent to {email}.",
        "invite": {
            "email": email,
            "token": str(invite.token),
            "expires_at": invite.expires_at.isoformat(),
            "link": link,
            "email_sent": email_ok,
        },
    })

@require_http_methods(["GET", "POST"])
def owner_accept(request):
    """
    Manager invite flow, simplified:
    - GET:
        If invite email has a verified phone -> send OTP + show code page.
        Else -> show phone form (and password if new user).
    - POST (from phone form):
        Normalize phone, validate password if needed, send OTP, then show code page.
    """

    # get invite
    token = request.GET.get("token") or request.POST.get("token") or ""
    invite = OwnerInvite.objects.filter(token=token).first()
    if not invite or not getattr(invite, "is_valid", False):
        return render(request, "core/owner_accept_invalid.html")

    restaurant_name = invite.restaurant.dba_name or invite.restaurant.legal_name

    user = User.objects.filter(email__iexact=invite.email).first()
    existing_user = bool(user)
    onfile_phone = find_verified_phone(user)

    # -------- GET --------
    if request.method == "GET":
        if onfile_phone:
            # auto-send OTP
            send_sms_otp(onfile_phone)
            request.session["mgr_accept"] = {
                "token": token,
                "email": invite.email.lower(),
                "phone": onfile_phone,
                "existing": existing_user,
            }
            return render(
                request,
                "core/owner_accept_code.html",
                {
                    "token": token,
                    "email": invite.email,
                    "phone_mask": mask(onfile_phone),
                    "restaurant_name": restaurant_name,
                },
            )
        else:
            return render(
                request,
                "core/owner_accept_phone.html",
                {
                    "token": token,
                    "email": invite.email,
                    "need_password": not existing_user,
                    "restaurant_name": restaurant_name,
                },
            )

    # -------- POST (phone form) --------
    phone_raw = (request.POST.get("phone") or "").strip()
    if not phone_raw:
        return JsonResponse({"ok": False, "error": "Phone is required."}, status=400)
    try:
        phone_e164 = to_e164_us(phone_raw)
    except Exception:
        return JsonResponse({"ok": False, "error": "Enter a valid US phone number."}, status=400)

    password = None
    if not existing_user:
        p1 = request.POST.get("password1") or ""
        p2 = request.POST.get("password2") or ""
        if not p1 or p1 != p2:
            return JsonResponse({"ok": False, "error": "Passwords didn't match."}, status=400)
        password = p1

    send_sms_otp(phone_e164)
    request.session["mgr_accept"] = {
        "token": token,
        "email": invite.email.lower(),
        "phone": phone_e164,
        "existing": existing_user,
        "password": password,
    }
    return render(
        request,
        "core/owner_accept_code.html",
        {
            "token": token,
            "email": invite.email,
            "phone_mask": mask(phone_e164),
            "restaurant_name": restaurant_name,
        },
    )


@require_POST
def owner_accept_verify(request):
    """Verify OTP, create user/profile if needed, accept invite, log in."""
    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "Bad JSON."}, status=400)

    stash = request.session.get("mgr_accept") or {}
    if not stash or stash.get("token") != data.get("token"):
        return JsonResponse({"ok": False, "error": "Session expired. Restart from invite link."}, status=400)

    code = (data.get("code") or "").strip()
    if not code:
        return JsonResponse({"ok": False, "error": "Missing code."}, status=400)

    try:
        if check_sms_otp(stash["phone"], code) != "approved":
            return JsonResponse({"ok": False, "error": "Invalid or expired code."}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Verification error: {e}"}, status=500)

    email = stash["email"]
    user = User.objects.filter(email__iexact=email).first()
    if not user:
        user = User.objects.create_user(username=email, email=email, password=stash.get("password"))
        user.is_active = True
        user.save()

    # create/attach manager profile
    mp, _ = OwnerProfile.objects.get_or_create(user=user)
    if not getattr(mp, "phone", None):
        mp.phone = stash["phone"]
    if hasattr(mp, "phone_verified"):
        mp.phone_verified = True
    if hasattr(mp, "email_verified"):
        mp.email_verified = True
    if not getattr(mp, "restaurant_id", None):
        invite = OwnerInvite.objects.filter(token=stash["token"]).first()
        if invite:
            mp.restaurant = invite.restaurant
            invite.accepted_at = timezone.now()
            invite.save(update_fields=["accepted_at"])
    mp.save()

    del request.session["mgr_accept"]
    login(request, user)
    return JsonResponse({"ok": True, "redirect": reverse("core:owner_dashboard")})