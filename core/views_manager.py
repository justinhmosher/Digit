from __future__ import annotations

import json
from datetime import timedelta, datetime

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpRequest, HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect
from django.views.decorators.http import require_GET, require_POST

from .models import ManagerProfile, RestaurantProfile, StaffProfile, TicketLink
from .omnivore import get_ticket, get_ticket_items
from django.apps import apps

_POSSIBLE_RATING_ATTRS = ("rating", "review_rating", "stars", "score")
_POSSIBLE_REVIEW_FIELDS = ("review_json", "raw_review_json", "raw_ticket_json", "extra_json")

def _to_star(v):
    try:
        f = float(v)
        # clamp to 0..5 and round to 3 decimals before casting if you prefer ints
        return max(0.0, min(5.0, f))
    except Exception:
        return None

def _dig_rating_from_mapping(m):
    if not isinstance(m, dict):
        return None

    # direct keys
    for k in _POSSIBLE_RATING_ATTRS:
        v = m.get(k)
        s = _to_star(v)
        if s is not None:
            return s

    # nested common spots
    for path in (
        ("review", "rating"),
        ("details", "rating"),
        ("feedback", "rating"),
        ("customer", "rating"),
        ("data", "rating"),
    ):
        cur = m
        try:
            for k in path:
                if not isinstance(cur, dict):
                    cur = {}
                cur = cur.get(k)
            s = _to_star(cur)
            if s is not None:
                return s
        except Exception:
            pass
    return None

def get_ticket_rating_from_anywhere(ticket_link):
    # --- 1) Review models
    for label in ("core.Review", "core.TicketReview", "reviews.Review"):
        try:
            Review = apps.get_model(label)
        except Exception:
            Review = None
        if not Review:
            continue

        r = None
        # FK by object
        try:
            r = Review.objects.filter(ticket_link=ticket_link).order_by("-id").first()
        except Exception:
            pass
        # Fallback: match by ticket_id (and restaurant if model has it)
        if r is None:
            try:
                filt = {"ticket_id": ticket_link.ticket_id}
                if hasattr(Review, "restaurant"):
                    filt["restaurant"] = ticket_link.restaurant
                r = Review.objects.filter(**filt).order_by("-id").first()
            except Exception:
                pass

        if r is not None:
            # scalar fields
            for attr in _POSSIBLE_RATING_ATTRS:
                if hasattr(r, attr):
                    s = _to_star(getattr(r, attr))
                    if s is not None:
                        return s
            # json/blob fields
            for blob_attr in _POSSIBLE_REVIEW_FIELDS:
                if hasattr(r, blob_attr):
                    s = _dig_rating_from_mapping(getattr(r, blob_attr))
                    if s is not None:
                        return s

    # --- 2) Direct attributes on TicketLink
    for attr in _POSSIBLE_RATING_ATTRS:
        if hasattr(ticket_link, attr):
            s = _to_star(getattr(ticket_link, attr))
            if s is not None:
                return s

    # --- 3) Dict/JSON blobs on TicketLink
    for blob_attr in _POSSIBLE_REVIEW_FIELDS:
        if hasattr(ticket_link, blob_attr):
            s = _dig_rating_from_mapping(getattr(ticket_link, blob_attr))
            if s is not None:
                return s

    return None

def _require_manager(request: HttpRequest):
    """Return (manager_profile, restaurant) or (None, None)."""
    mp = getattr(request.user, "manager_profile", None)
    if mp is None:
        mp = (
            ManagerProfile.objects
            .select_related("restaurant")
            .filter(user=request.user)
            .first()
        )
    if not mp or not getattr(mp, "restaurant", None):
        return None, None
    return mp, mp.restaurant


@login_required
def manager_dashboard(request: HttpRequest) -> HttpResponse:
    mp, restaurant = _require_manager(request)
    if not mp:
        return redirect("/restaurant/signin?tab=manager")

    if restaurant and request.session.get("current_restaurant_id") != restaurant.id:
        request.session["current_restaurant_id"] = restaurant.id
        request.session.modified = True

    return render(request, "core/manager_dashboard.html", {"mp": mp, "restaurant": restaurant})


from datetime import datetime, time, timedelta
from django.utils.dateparse import parse_date
from django.db.models import Q

@ensure_csrf_cookie
@require_GET
@login_required
def manager_api_state(request: HttpRequest) -> JsonResponse:
    """
    Dashboard data for manager:
      - staff list (name only + ids for Remove)
      - open tickets summary
      - recent closed tickets within optional date range
    """
    mp, rp = _require_manager(request)
    if not mp or not rp:
        return JsonResponse({"ok": False, "error": "Not a manager for any restaurant."}, status=403)

    q = (request.GET.get("q") or "").strip().lower()
    start = (request.GET.get("start") or "").strip()
    end   = (request.GET.get("end") or "").strip()

    # Staff (derive a simple name from the user)
    staff = []
    staff_qs = (
        StaffProfile.objects
        .select_related("user")
        .filter(restaurant=rp)
        .order_by("user__email")
    )
    for s in staff_qs:
        u = getattr(s, "user", None)
        email = getattr(u, "email", "") if u else ""
        name = (u.get_full_name() if u else "") or (email.split("@")[0] if email else "")
        staff.append({"id": s.id, "name": name})

    # Open tickets, group by ticket id
    open_map = {}
    open_qs = (
        TicketLink.objects
        .select_related("member")
        .filter(restaurant=rp, status="open")
        .order_by("-opened_at")[:400]
    )
    for tl in open_qs:
        entry = open_map.setdefault(tl.ticket_id, {
            "ticket_id": tl.ticket_id,
            "ticket_number": tl.ticket_number or None,
            "server": tl.server_name or "",
            "members": [],
            "due_cents": 0,
        })
        entry["members"].append(tl.member.number if tl.member else "")
        entry["due_cents"] = max(entry["due_cents"], tl.last_total_cents or 0)
    open_list = list(open_map.values())

    # Date window for "recent"
    recent_qs = TicketLink.objects.select_related("member").filter(restaurant=rp, status="closed")
    if start:
        try:
            recent_qs = recent_qs.filter(closed_at__date__gte=start)
        except Exception:
            pass
    if end:
        try:
            recent_qs = recent_qs.filter(closed_at__date__lte=end)
        except Exception:
            pass

    recent = []
    for tl in recent_qs.order_by("-closed_at")[:500]:
        row = {
            "ticket_id": tl.ticket_id,
            "ticket_number": tl.ticket_number or None,
            "member": tl.member.number if tl.member else "",
            "server": tl.server_name or "",
            "closed_at": tl.closed_at.strftime("%Y-%m-%d %H:%M") if tl.closed_at else "",
            "total_cents": (tl.paid_cents or tl.total_cents or tl.last_total_cents or 0),
            "ticket_link_id": tl.id,  # <-- NEW: used by the UI "Review" button
        }
        if q:
            hay = f'{row["ticket_number"] or ""} {row["member"] or ""}'.lower()
            if q not in hay:
                continue
        recent.append(row)

    return JsonResponse({"ok": True, "staff": staff, "open": open_list, "recent": recent})




@csrf_protect
@require_POST
@login_required
def manager_api_remove_staff(request: HttpRequest) -> JsonResponse:
    mp, rp = _require_manager(request)
    if not mp or not rp:
        return JsonResponse({"ok": False, "error": "Not authorized."}, status=403)

    try:
        data = json.loads(request.body.decode() or "{}")
    except Exception:
        data = {}
    sid = data.get("staff_id")
    if not sid:
        return JsonResponse({"ok": False, "error": "Missing staff_id."}, status=400)

    sp = StaffProfile.objects.filter(id=sid, restaurant=rp).first()
    if not sp:
        return JsonResponse({"ok": False, "error": "Staff not found."}, status=404)

    sp.restaurant = None  # unlink access
    sp.save(update_fields=["restaurant"])

    return JsonResponse({"ok": True})



# DROP-IN: replace the whole function in core/views_manager.py

@require_GET
@login_required
def manager_api_ticket_detail(request: HttpRequest, ticket_id: str) -> JsonResponse:
    mp, rp = _require_manager(request)
    if not mp or not rp:
        return JsonResponse({"ok": False, "error": "Not authorized."}, status=403)

    tl = (
        TicketLink.objects
        .filter(restaurant=rp, ticket_id=str(ticket_id))
        .order_by("-opened_at")
        .first()
    )
    if not tl:
        return JsonResponse({"ok": False, "error": "Ticket not found."}, status=404)

    # ---------- OPEN: pull live from POS; expose unit and line totals ----------
    if tl.status == "open" and (rp.omnivore_location_id or "").strip():
        try:
            t = get_ticket(rp.omnivore_location_id, tl.ticket_id)
            items = get_ticket_items(rp.omnivore_location_id, tl.ticket_id)
        except Exception as e:
            return JsonResponse({"ok": False, "error": f"POS error: {e}"}, status=502)

        rows = []
        for it in items:
            qty  = int(it.get("quantity", 1) or 1)
            unit = int(it.get("price", 0) or 0)  # Omnivore price is per-unit cents
            rows.append({
                "name": it.get("name"),
                "qty": qty,
                "unit_cents": unit,
                "line_total_cents": unit * qty,
                "cents": unit,  # back-compat for older UIs
            })

        totals = (t or {}).get("totals") or {}
        tax = int(totals.get("tax", 0) or 0)
        tip = int(totals.get("tip", 0) or 0)

        try:
            due_val = totals.get("due")
            total = int(due_val) if due_val is not None else int(totals.get("total", 0) or 0)
        except Exception:
            total = 0
        if total <= 0:
            total = sum(r["line_total_cents"] for r in rows) + tax + tip

        # Display rule for OPEN tickets: Subtotal == Total
        return JsonResponse({
            "ok": True,
            "ticket_id": tl.ticket_id,
            "ticket_number": tl.ticket_number or t.get("ticket_number") or t.get("number") or tl.ticket_id,
            "member": tl.member.number if tl.member else "",
            "server": tl.server_name or ((t.get("_embedded") or {}).get("employee") or {}).get("check_name", ""),
            "items": rows,
            "subtotal_cents": total,
            "tax_cents": tax,
            "tip_cents": tip,
            "total_cents": total,
            "is_open": True,
        })

    # ---------- CLOSED: use saved snapshot; expose unit and line totals ----------
    rows = []
    for it in (tl.items_json or []):
        name  = it.get("name") or it.get("label") or "Item"
        qty   = int(it.get("qty") or it.get("quantity") or 1)
        # Attempt multiple fields for unit; coerce to cents
        unit  = int(it.get("price_cents") or it.get("unit_cents") or it.get("cents") or it.get("price") or 0)
        line  = int(it.get("total_cents") or it.get("line_total_cents") or (unit * qty))
        rows.append({
            "name": name,
            "qty": qty,
            "unit_cents": unit,
            "line_total_cents": line,
            "cents": unit,  # back-compat
        })

    return JsonResponse({
        "ok": True,
        "ticket_id": tl.ticket_id,
        "ticket_number": tl.ticket_number or tl.ticket_id,
        "member": tl.member.number if tl.member else "",
        "server": tl.server_name or "",
        # CLOSED mapping: Subtotal = total_cents, Total = paid_cents (if present)
        "items": rows,
        "subtotal_cents": int(tl.total_cents or 0),
        "tax_cents": int(tl.tax_cents or 0),
        "tip_cents": int(tl.tip_cents or 0),
        "total_cents": int(tl.paid_cents or (tl.total_cents or 0) + (tl.tax_cents or 0) + (tl.tip_cents or 0)),
        "is_open": False,
    })


from io import BytesIO
from datetime import datetime, time, timedelta
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

@require_GET
@login_required
def manager_export(request: HttpRequest) -> HttpResponse:
    mp, rp = _require_manager(request)
    if not mp or not rp:
        return HttpResponse("Not authorized.", status=403)

    q = (request.GET.get("q") or "").strip().lower()
    start = (request.GET.get("start") or "").strip()
    end   = (request.GET.get("end") or "").strip()

    qs = (
        TicketLink.objects
        .select_related("member")
        .filter(restaurant=rp, status="closed")
    )
    if start:
        try:
            qs = qs.filter(closed_at__date__gte=start)
        except Exception:
            pass
    if end:
        try:
            qs = qs.filter(closed_at__date__lte=end)
        except Exception:
            pass

    rows = []
    for tl in qs.order_by("closed_at"):
        # optional search filter
        if q:
            hay = f"{(tl.ticket_number or tl.ticket_id)} {(tl.member.number if tl.member else '')}".lower()
            if q not in hay:
                continue

        subtotal = int(tl.total_cents or 0)   # <-- per your mapping
        tax      = int(tl.tax_cents or 0)
        tip      = int(tl.tip_cents or 0)
        total    = int(tl.paid_cents or (subtotal + tax + tip))  # <-- Total shows what was paid

        rows.append({
            "closed": tl.closed_at,
            "ticket": tl.ticket_number or tl.ticket_id,
            "member": tl.member.number if tl.member else "",
            "server": tl.server_name or "",
            "subtotal": subtotal,
            "tax": tax,
            "tip": tip,
            "total": total,
            "pos_ref": tl.pos_ref or "",
        })

    # Build workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Closed"

    headers = ["Closed", "Ticket", "Member", "Server", "Subtotal", "Tax", "Tip", "Total", "POS Ref"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    currency_fmt = u'[$$-409]#,##0.00'

    for r in rows:
        ws.append([
            r["closed"].strftime("%Y-%m-%d %H:%M") if r["closed"] else "",
            r["ticket"],
            r["member"],
            r["server"],
            r["subtotal"]/100.0,
            r["tax"]/100.0,
            r["tip"]/100.0,
            r["total"]/100.0,
            r["pos_ref"],
        ])

    # Format currency columns E..H
    if rows:
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=8):
            for cell in row:
                cell.number_format = currency_fmt

    # Column widths
    widths = [18, 12, 14, 12, 12, 12, 12, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Grand total row (sum of Total column)
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=7, value="Grand total").font = Font(bold=True)
    total_col = get_column_letter(8)
    ws.cell(row=total_row, column=8, value=f"=SUM({total_col}2:{total_col}{total_row-1})").number_format = currency_fmt
    ws.cell(row=total_row, column=8).font = Font(bold=True)

    # Return file
    rest_name = getattr(rp, "dba_name", "") or getattr(rp, "legal_name", "") or "restaurant"
    filename = f"{rest_name.replace(' ', '_')}_closed_{timezone.now().strftime('%Y%m%d')}.xlsx"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def get_ticket_rating_from_anywhere(ticket_link):
    # 1) Review models
    for label in ("core.Review", "core.TicketReview", "reviews.Review"):
        try:
            Review = apps.get_model(label)
        except Exception:
            Review = None
        if not Review:
            continue

        try:
            r = Review.objects.filter(ticket_link=ticket_link).order_by("-id").first()
        except Exception:
            r = None

        if r is None:
            try:
                filt = {"ticket_id": ticket_link.ticket_id}
                if hasattr(Review, "restaurant"):
                    filt["restaurant"] = ticket_link.restaurant
                r = Review.objects.filter(**filt).order_by("-id").first()
            except Exception:
                r = None

        if r is not None:
            for attr in _POSSIBLE_RATING_ATTRS:
                try:
                    val = getattr(r, attr, None)
                    if val is not None:
                        val = int(val)
                        if 0 <= val <= 5:
                            return val
                except Exception:
                    pass
            for blob_attr in _POSSIBLE_REVIEW_FIELDS:
                try:
                    blob = getattr(r, blob_attr, None)
                    if blob:
                        v = _dig_rating_from_mapping(blob)
                        if v is not None:
                            return v
                except Exception:
                    pass

    # 2) Direct attrs on TicketLink
    for attr in _POSSIBLE_RATING_ATTRS:
        try:
            val = getattr(ticket_link, attr, None)
            if val is not None:
                val = int(val)
                if 0 <= val <= 5:
                    return val
        except Exception:
            pass

    # 3) Dict/JSON blobs on TicketLink
    for blob_attr in _POSSIBLE_REVIEW_FIELDS:
        try:
            blob = getattr(ticket_link, blob_attr, None)
            if blob:
                v = _dig_rating_from_mapping(blob)
                if v is not None:
                    return v
        except Exception:
            pass

    return None

@login_required
@require_GET
def manager_api_menu_item_ratings(request: HttpRequest) -> JsonResponse:
    """
    Mirrors owner_api_menu_item_ratings but scoped to the manager's single restaurant.
    Response shape:
      { ok: True, items: [ {menu_item_id, name, category, price_cents, avg_rating, num_rated_tickets, total_qty_on_rated_tickets, num_all_tickets, total_qty_all_tickets} ... ] }
    """
    mp, rp = _require_manager(request)
    if not mp or not rp:
        return JsonResponse({"ok": False, "error": "Not authorized."}, status=403)

    start_s = (request.GET.get("start") or "").strip()
    end_s   = (request.GET.get("end") or "").strip()

    qs = TicketLink.objects.filter(restaurant=rp, status="closed")
    if start_s:
        try: qs = qs.filter(closed_at__date__gte=parse_date(start_s))
        except Exception: pass
    if end_s:
        try: qs = qs.filter(closed_at__date__lte=parse_date(end_s))
        except Exception: pass

    # Menu metadata from RestaurantProfile.menu_cache, if present
    item_meta = { str(x.get("id")): x for x in (getattr(rp, "menu_cache", None) or []) }

    agg = {}
    for tl in qs.iterator():
        rating = get_ticket_rating_from_anywhere(tl)  # may be None

        for row in (tl.items_json or []):
            mid = str(row.get("menu_item_id") or row.get("id") or "").strip()
            name = (row.get("name") or row.get("label") or "").strip() or "Unknown item"
            key = mid or f"name:{name}"

            try:
                qty = int(row.get("quantity") or row.get("qty") or 1)
            except Exception:
                qty = 1

            rec = agg.setdefault(key, {
                "menu_item_id": mid or None,
                "name": name,
                "category": "",
                "price_cents": 0,
                "sum": 0,
                "n": 0,
                "qty_rated_tickets": 0,
                "qty_all_tickets": 0,
                "num_rated_tickets": 0,
                "num_all_tickets": 0,
            })

            # enrich from cache
            if mid and mid in item_meta:
                meta = item_meta[mid]
                if not rec["category"]:
                    rec["category"] = meta.get("category") or ""
                if not rec["price_cents"]:
                    try:
                        rec["price_cents"] = int(meta.get("price_cents") or 0)
                    except Exception:
                        pass

            # fallback to ticket unit price if cache has none
            if not rec["price_cents"]:
                for k in ("unit_cents", "price_cents", "cents", "unit_price", "price"):
                    v = row.get(k)
                    if v is None:
                        continue
                    try:
                        iv = int(v)
                        rec["price_cents"] = iv if iv >= 100 else iv * 100
                        break
                    except Exception:
                        try:
                            from decimal import Decimal
                            rec["price_cents"] = int(Decimal(str(v)) * 100)
                            break
                        except Exception:
                            pass

            rec["qty_all_tickets"] += qty
            rec["num_all_tickets"] += 1

            if rating is not None:
                rec["sum"] += int(rating)
                rec["n"]   += 1
                rec["qty_rated_tickets"] += qty
                rec["num_rated_tickets"] += 1

    out = []
    for _, r in agg.items():
        avg = (r["sum"] / r["n"]) if r["n"] else None
        out.append({
            "menu_item_id": r["menu_item_id"] or "",
            "name": r["name"],
            "category": r["category"],
            "price_cents": int(r["price_cents"] or 0),
            "avg_rating": round(avg, 3) if avg is not None else None,
            "num_rated_tickets": r["num_rated_tickets"],
            "total_qty_on_rated_tickets": r["qty_rated_tickets"],
            # extras to match owner endpoint exactly
            "num_all_tickets": r["num_all_tickets"],
            "total_qty_all_tickets": r["qty_all_tickets"],
        })

    def sort_key(row):
        rated = row["avg_rating"] is not None
        return (0 if rated else 1, -(row["avg_rating"] or 0), -row.get("num_all_tickets", 0))

    out.sort(key=sort_key)
    return JsonResponse({"ok": True, "items": out})

@login_required
@require_GET
def manager_api_staff_ratings(request: HttpRequest) -> JsonResponse:
    """
    Mirrors owner_api_staff_ratings; includes an 'ALL' row.
    Response shape:
      { ok: True, synced_at, staff: [ {staff_key, name, avg_rating, num_rated_tickets, is_active_in_pos, num_all_tickets}, ... ] }
    """
    mp, rp = _require_manager(request)
    if not mp or not rp:
        return JsonResponse({"ok": False, "error": "Not authorized."}, status=403)

    start_s = (request.GET.get("start") or "").strip()
    end_s   = (request.GET.get("end") or "").strip()
    qs = TicketLink.objects.filter(restaurant=rp, status="closed")
    if start_s:
        try: qs = qs.filter(closed_at__date__gte=parse_date(start_s))
        except Exception: pass
    if end_s:
        try: qs = qs.filter(closed_at__date__lte=parse_date(end_s))
        except Exception: pass

    staff_cache = getattr(rp, "staff_cache", None) or []
    by_ck_lower = { (s.get("check_name") or "").strip().lower(): s for s in staff_cache if (s.get("check_name") or "").strip() }
    by_nm_lower = { (s.get("name") or "").strip().lower():       s for s in staff_cache if (s.get("name") or "").strip() }

    agg = {}
    def seed_row(key, display, active=True):
        if key not in agg:
            agg[key] = {"display": display or "", "active": bool(active), "sum": 0, "n": 0, "tickets_all": 0, "tickets_rated": 0}

    # Seed cached staff so they appear even with 0 tickets (but only if we have a display)
    for s in staff_cache:
        key = (str(s.get("id") or "").strip()
               or (s.get("check_name") or "").strip()
               or (s.get("name") or "").strip()
               or f"seed:{id(s)}")
        disp = (s.get("check_name") or s.get("name") or "").strip()
        if disp:
            seed_row(key, disp, s.get("is_active", True))

    agg_all = {"display": "All staff", "active": True, "sum": 0, "n": 0, "tickets_all": 0, "tickets_rated": 0}

    for tl in qs.iterator():
        rating = get_ticket_rating_from_anywhere(tl)

        # ALL
        agg_all["tickets_all"] += 1
        if rating is not None:
            agg_all["sum"] += int(rating)
            agg_all["n"]   += 1
            agg_all["tickets_rated"] += 1

        resolved = (tl.server_name or "").strip()
        key = None
        display = None

        if resolved:
            low = resolved.lower()
            if low in by_ck_lower:
                s = by_ck_lower[low]
                key = str(s.get("id") or s.get("check_name") or resolved)
                display = (s.get("check_name") or s.get("name") or resolved)
                seed_row(key, display, s.get("is_active", True))
            elif low in by_nm_lower:
                s = by_nm_lower[low]
                key = str(s.get("id") or s.get("name") or resolved)
                display = (s.get("check_name") or s.get("name") or resolved)
                seed_row(key, display, s.get("is_active", True))
            else:
                display = resolved
                key = resolved
                if display:
                    seed_row(key, display, True)

        if not key:
            try:
                raw = tl.raw_ticket_json or {}
                emb = (raw.get("_embedded") or {})
                emp = emb.get("employee") or raw.get("employee") or {}
                pos_name = (emp.get("check_name") or emp.get("name") or "").strip()
            except Exception:
                pos_name = ""
            if pos_name:
                low = pos_name.lower()
                if low in by_ck_lower:
                    s = by_ck_lower[low]
                    key = str(s.get("id") or s.get("check_name") or pos_name)
                    display = (s.get("check_name") or s.get("name") or pos_name)
                    seed_row(key, display, s.get("is_active", True))
                elif low in by_nm_lower:
                    s = by_nm_lower[low]
                    key = str(s.get("id") or s.get("name") or pos_name)
                    display = (s.get("check_name") or s.get("name") or pos_name)
                    seed_row(key, display, s.get("is_active", True))
                else:
                    display = pos_name
                    key = pos_name
                    if display:
                        seed_row(key, display, True)

        if not key:
            continue

        row = agg[key]
        row["tickets_all"] += 1
        if rating is not None:
            row["sum"] += int(rating)
            row["n"]   += 1
            row["tickets_rated"] += 1

    # Shape output exactly like owner endpoint
    out = []
    avg_all = (agg_all["sum"] / agg_all["n"]) if agg_all["n"] else None
    out.append({
        "staff_key": "ALL",
        "name": agg_all["display"],
        "avg_rating": round(avg_all, 3) if avg_all is not None else None,
        "num_rated_tickets": agg_all["tickets_rated"],
        "is_active_in_pos": True,
        "num_all_tickets": agg_all["tickets_all"],
    })
    for key, r in agg.items():
        avg = (r["sum"] / r["n"]) if r["n"] else None
        out.append({
            "staff_key": key,
            "name": r["display"],
            "avg_rating": round(avg, 3) if avg is not None else None,
            "num_rated_tickets": r["tickets_rated"],
            "is_active_in_pos": bool(r["active"]),
            "num_all_tickets": r["tickets_all"],
        })

    def sort_key(row):
        if row["staff_key"] == "ALL": return (-999, 0, 0)
        rated = row["avg_rating"] is not None
        return (0 if rated else 1, -(row["avg_rating"] or 0), -row["num_all_tickets"])

    out[1:] = sorted(out[1:], key=sort_key)

    return JsonResponse({
        "ok": True,
        "synced_at": (rp.staff_cache_synced_at.isoformat() if getattr(rp, "staff_cache_synced_at", None) else None),
        "staff": out,
    })
def _manager_restaurant_or_404(request: HttpRequest):
    mp, rp = _require_manager(request)
    if not mp or not rp:
        raise Http404("Not a manager for any restaurant")
    return rp

@login_required
@require_GET
def manager_ticket_review_json(request: HttpRequest, ticket_link_id: int) -> JsonResponse:
    """
    Returns a normalized review payload for a given TicketLink id, mirroring the owner view shape:
      {
        ok, ticket_id, ticket_number, member,
        has_review: bool,
        review: { id, stars, rating, comment, created_at, reviewer_name, reviewer_email } | null
      }
    """
    rp = _manager_restaurant_or_404(request)
    try:
        tl = TicketLink.objects.select_related("member").get(pk=ticket_link_id, restaurant=rp)
    except TicketLink.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Ticket not found."}, status=404)

    # Try to locate a Review model we may have
    ReviewModel = None
    for label in ("core.Review", "core.TicketReview", "reviews.Review"):
        try:
            ReviewModel = apps.get_model(label)
            break
        except Exception:
            continue

    review = None
    if ReviewModel:
        # Prefer FK by object, fall back to ticket_id (+restaurant if available)
        try:
            review = ReviewModel.objects.filter(ticket_link=tl).order_by("-created_at", "-id").first()
        except Exception:
            review = None
        if review is None:
            try:
                filt = {"ticket_id": tl.ticket_id}
                if hasattr(ReviewModel, "restaurant"):
                    filt["restaurant"] = tl.restaurant
                review = ReviewModel.objects.filter(**filt).order_by("-created_at", "-id").first()
            except Exception:
                review = None

    out = {
        "ok": True,
        "ticket_id": tl.ticket_id,
        "ticket_number": tl.ticket_number,
        "member": getattr(tl.member, "member_id", None) or getattr(tl.member, "id", None),
        "has_review": bool(review),
        "review": None,
    }

    if review:
        # Normalize stars 0..5 from scalar attrs or blobs
        stars = None
        for attr in ("stars", "rating", "score"):
            try:
                v = getattr(review, attr, None)
                if v is not None:
                    stars = max(0, min(5, int(v)))
                    break
            except Exception:
                pass
        if stars is None:
            for blob_attr in ("review_json", "raw_review_json", "raw_ticket_json", "extra_json"):
                try:
                    blob = getattr(review, blob_attr, None)
                    if blob and isinstance(blob, dict):
                        v = (blob.get("rating")
                             or (blob.get("review") or {}).get("rating")
                             or (blob.get("details") or {}).get("rating"))
                        if v is not None:
                            stars = max(0, min(5, int(v)))
                            break
                except Exception:
                    pass
        stars = stars if stars is not None else 0

        out["review"] = {
            "id": review.id,
            "stars": stars,
            "rating": stars,
            "comment": getattr(review, "comment", "") or "",
            "created_at": getattr(review, "created_at", None).isoformat() if getattr(review, "created_at", None) else None,
            "reviewer_name": getattr(review, "reviewer_name", "") or "",
            "reviewer_email": getattr(review, "reviewer_email", "") or "",
        }

    return JsonResponse(out)
